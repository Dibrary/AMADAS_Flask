from abc import *
import xlrd
from lxml import etree
from openpyxl.styles import *
from openpyxl.cell import Cell
from openpyxl.chart import (PieChart, LineChart, ProjectedPieChart, Reference)
from openpyxl.chart.series import DataPoint
from pylab import title, figure, xlabel, ylabel, xticks, bar, legend, axis, savefig
from matplotlib import pyplot as plt
import matplotlib
import numpy as np
from decimal import *


class Report(metaclass=ABCMeta):
# report용 추상클래스. (이 추상클래스 밑으로, all_kpi_report, multi_validation_report, validation_report, event_report가 상속받아 사용한다)

    def file_path(self): # 파일 생성할 path를 xml파일에 설정된 값으로 가져온다.
        root = (etree.parse("opc_modbus_db_ip.xml")).getroot()
        for i in root.iter("report"):
            path = i.findtext("path")
        return str(path)

    def design_for_report(self): # 기본 디자인 정보는 어떤 report나 같기 때문에, 이렇게 하나로 만들고, 반환하게함.
        headerFill = PatternFill(start_color='FFFFFFCC', end_color="FFFFFFCC", fill_type="solid")
        header2Fill = PatternFill(start_color='FFFFCC00', end_color="FFFFCC00", fill_type="solid")
        tableFill = PatternFill(start_color='FFCCCCCC', end_color="FFCCCCCC", fill_type="solid")

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        top_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        return (headerFill, header2Fill, tableFill, thin_border, top_bottom_border)

    def cell_method(self, ws, cell, text, size, bold=None):  # cell을 설정한 것에 맞게 적용한 후에 반환하는 기능
        ws[cell] = text
        if bold != None:
            ws[cell].font = Font(name="맑은 고딕", size=size, bold=True)
        else:
            ws[cell].font = Font(name="맑은 고딕", size=size)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        return ws

    def title_cell(self, ws, div, where):  # 단순히 title을 div에 따라 분류하는 기능
        if where == "TITLE":
            if self.div == "Validation":
                ws = self.cell_method(ws, 'A1', 'ANALYZER \n VALIDATION REPORT', 18, True)
            elif self.div == "Trend":
                ws = self.cell_method(ws, 'A1', 'ANALYZER \n TREND REPORT', 18, True)
            elif self.div == "ALL":
                ws = self.cell_method(ws, 'A1', 'ANALYZER \n ALL KPI REPORT', 18, True)
            elif self.div == "Alarm":
                ws = self.cell_method(ws, 'A1', 'ANALYZER \n ALARM REPORT', 18, True)
            elif self.div == "Maintenance":
                ws = self.cell_method(ws, 'A1', 'ANALYZER \n MAINTENANCE REPORT', 18, True)
        elif where == "VALUE":
            if self.div == "Validation":
                ws = self.cell_method(ws, 'A32', 'ANALYZER VALIDATION DATA', 16, True)
            elif self.div == "Trend":
                ws = self.cell_method(ws, 'A32', 'ANALYZER TREND DATA', 16, True)
            elif self.div == "ALL":
                ws = self.cell_method(ws, 'A1', 'ANALYZER \n ALL KPI REPORT', 18, True)
        return ws

    @abstractmethod
    def kpi_chart(self, ws):
        pass

    @abstractmethod
    def trend_chart(self, ws, length):
        pass

    @abstractmethod
    def basic_inform_cell(self, ws, start_idx, end_idx, title, value, style):
        pass

    @abstractmethod
    def column_cell(self, ws, div, style):
        pass

    @abstractmethod
    def data_preprocess(self, data, div):
        pass

    @abstractmethod
    def data_array(self, ws, div, datas, linestyle, fillstyle):
        pass

    @abstractmethod
    def make_excel_report(self):
        pass