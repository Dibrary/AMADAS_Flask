import sys, time, datetime, openpyxl
from lxml import etree

import numpy as np
from matplotlib import pyplot as plt
import matplotlib
import xlrd
from openpyxl.styles import *
from openpyxl.cell import Cell
from openpyxl.chart import (PieChart, LineChart, ProjectedPieChart, Reference)
from openpyxl.chart.series import DataPoint
from decimal import *
from pylab import title, figure, xlabel, ylabel, xticks, bar, legend, axis, savefig

from db_module import *
from kpi_module import *
from report_abstract import *


class Validation_Report(Report):
    ana_tag: str
    user_id: str
    start_dt: str
    end_dt: str
    file_name: str
    div: str  # 어떤 종류의 report인지 분류
    bottle_tag: str
    dbmodule: object
    kpimodule: object

    def __init__(self, ana_tag, user_id, start_dt, end_dt, file_name, div, bottle_tag=None):
        self.ana_tag = ana_tag
        self.user_id = user_id
        self.start_dt = start_dt
        self.end_dt = end_dt
        self.file_name = file_name
        self.div = div
        self.bottle_tag = bottle_tag
        self.dbmodule = withDB()
        self.kpimodule = KPI_module(start_dt, end_dt, div, ana_tag)

    def basic_inform_cell(self, ws: object, start_idx: int, end_idx: int, title: list, value: list, style: object):
        for i in range(start_idx, end_idx):
            ws.merge_cells('E' + str(i) + ':F' + str(i))
            ws = self.cell_method(ws, 'E' + str(i), title[i - start_idx], 11, True)
            ws.merge_cells('G' + str(i) + ':H' + str(i))
            ws = self.cell_method(ws, 'G' + str(i), value[i - start_idx], 11)
            ws['E' + str(i)].border, ws['G' + str(i)].border, ws['F' + str(i)].border, ws[
                'H' + str(i)].border = style, style, style, style
        return ws

    def column_cell(self, ws: object, div: str, style: object):
        if div == "Validation" or div == 'Trend':
            if div == 'Trend':
                columns = ['A', 'B', 'D', 'F', 'H']
                column_name = ['no', 'Tag', 'Process', 'Check Date', 'etc']
                width_no = [5, 10, 13, 13, 10]
                ws.merge_cells('B34:C34')
                ws.merge_cells('D34:E34')
                ws.merge_cells('F34:G34')
                for i in columns:
                    ws[i + '34'] = column_name[columns.index(i)]
                    ws[i + '34'].font = Font(name="맑은 고딕", size=11, bold=True)
                    ws[i + '34'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[i + '34'].border = style
                    ws.column_dimensions[i].width = width_no[columns.index(i)]
                ws['G34'].border, ws['C34'].border, ws['E34'].border = style, style, style
            elif div == 'Validation':
                columns = ['A', 'B', 'C', 'D', 'E', 'F', 'H']
                column_name = ['no', 'Tag', 'Type', 'Validation', 'Start Date', 'Check Date', 'Result']
                width_no = [5, 15, 10, 15, 15, 7, 10]
                ws.merge_cells('F34:G34')
                for i in columns:
                    ws[i + '34'] = column_name[columns.index(i)]
                    ws[i + '34'].font = Font(name="맑은 고딕", size=11, bold=True)
                    ws[i + '34'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[i + '34'].border = style
                    ws.column_dimensions[i].width = width_no[columns.index(i)]
                ws['G34'].border = style
        return ws

    def data_preprocess(self, data: list, div: str):
        if div == "Validation":  # validation의 경우,
            tag, v_type, validation, start_time, check_time, result, user_id = [], [], [], [], [], [], []
            for k in range(0, len(data)):
                tag.append(data[k][0])
                v_type.append(data[k][1])
                validation.append(round(float(data[k][2]), 2))
                start_time.append(str(data[k][3])[2:16])  # 분 까지만 나타낸다.
                check_time.append(str(data[k][4])[2:16])
                result.append(data[k][5])
                user_id.append(data[k][6])
            return (tag, v_type, validation, start_time, check_time, result, user_id)
        elif div == "Trend":
            tag, process, check_time = [], [], []
            for i in range(0, len(data)):
                tag.append(data[i][1])
                process.append(round(float(data[i][2]), 2))
                check_time.append(str(data[i][3])[2:16])
            return (tag, process, check_time)

    def data_array(self, ws: object, div: str, datas: list, linestyle: object, fillstyle: object):
        if div == "Validation":
            result_data = self.data_preprocess(datas, div)
            tag, v_type, value, start_dt, check_dt, results, user_id = result_data[0], result_data[1], result_data[2], \
                                                                       result_data[3], result_data[4], result_data[5], \
                                                                       result_data[6]
            row = len(value)
            for i in range(35, 35 + row):
                ws.merge_cells('F' + str(i) + ':G' + str(i))
                ws['F' + str(i)].font = Font(name="맑은 고딕", size=11)
                ws['F' + str(i)].alignment = Alignment(horizontal='center', vertical='center')
                ws['F' + str(i)].border, ws['G' + str(i)].border, ws[
                    'H' + str(i)].border = linestyle, linestyle, linestyle
                if i % 2 == 0:
                    ws = self.cell_method(ws, 'A' + str(i), str(i - 35), 11)  # no
                    ws = self.cell_method(ws, 'B' + str(i), tag[i - 35], 11)  # validation_tag
                    ws = self.cell_method(ws, 'C' + str(i), v_type[i - 35], 11)  # validation_type
                    ws = self.cell_method(ws, 'D' + str(i), value[i - 35], 11)  # validation_value
                    ws = self.cell_method(ws, 'E' + str(i), start_dt[i - 35], 11)  # start_datetime
                    ws = self.cell_method(ws, 'F' + str(i), check_dt[i - 35], 11)  # check_datetime
                    ws = self.cell_method(ws, 'H' + str(i), results[i - 35], 11)  # result
                    ws['A' + str(i)].fill, ws['B' + str(i)].fill, ws['C' + str(i)].fill, ws['D' + str(i)].fill, ws[
                        'E' + str(i)].fill, ws['F' + str(i)].fill, ws['G' + str(i)].fill, ws['H' + str(
                        i)].fill = fillstyle, fillstyle, fillstyle, fillstyle, fillstyle, fillstyle, fillstyle, fillstyle
                else:
                    ws = self.cell_method(ws, 'A' + str(i), str(i - 35), 11)  # no
                    ws = self.cell_method(ws, 'B' + str(i), tag[i - 35], 11)  # validation_tag
                    ws = self.cell_method(ws, 'C' + str(i), v_type[i - 35], 11)  # validation_type
                    ws = self.cell_method(ws, 'D' + str(i), value[i - 35], 11)  # validation_value
                    ws = self.cell_method(ws, 'E' + str(i), start_dt[i - 35], 11)  # start_datetime
                    ws = self.cell_method(ws, 'F' + str(i), check_dt[i - 35], 11)  # check_datetime
                    ws = self.cell_method(ws, 'H' + str(i), results[i - 35], 11)  # result
                ws['A' + str(i)].border, ws['B' + str(i)].border, ws['C' + str(i)].border, ws['D' + str(i)].border, ws[
                    'E' + str(i)].border = linestyle, linestyle, linestyle, linestyle, linestyle
        elif div == "Trend":
            result_data = self.data_preprocess(datas, div)
            tag, value, check_dt = result_data[0], result_data[1], result_data[2]
            row = len(value)
            for i in range(35, 35 + row):
                ws.merge_cells('B' + str(i) + ':C' + str(i))
                ws.merge_cells('D' + str(i) + ':E' + str(i))
                ws.merge_cells('F' + str(i) + ':G' + str(i))
                ws['F' + str(i)].font = Font(name="맑은 고딕", size=11)
                ws['F' + str(i)].alignment = Alignment(horizontal='center', vertical='center')
                ws['F' + str(i)].border, ws['G' + str(i)].border = linestyle, linestyle
                if i % 2 == 0:
                    ws = self.cell_method(ws, 'A' + str(i), str(i - 35), 11)  # no
                    ws = self.cell_method(ws, 'B' + str(i), tag[i - 35], 11)  # validation_tag
                    ws = self.cell_method(ws, 'D' + str(i), value[i - 35], 11)  # process_value
                    ws = self.cell_method(ws, 'F' + str(i), check_dt[i - 35], 11)  # check_datetime
                    ws['A' + str(i)].fill, ws['B' + str(i)].fill, ws['C' + str(i)].fill, ws['D' + str(i)].fill, ws[
                        'E' + str(i)].fill, ws['F' + str(i)].fill, ws[
                        'G' + str(i)].fill = fillstyle, fillstyle, fillstyle, fillstyle, fillstyle, fillstyle, fillstyle
                else:
                    ws = self.cell_method(ws, 'A' + str(i), str(i - 35), 11)  # no
                    ws = self.cell_method(ws, 'B' + str(i), tag[i - 35], 11)  # validation_tag
                    ws = self.cell_method(ws, 'D' + str(i), value[i - 35], 11)  # process_value
                    ws = self.cell_method(ws, 'F' + str(i), check_dt[i - 35], 11)  # check_datetime
                ws['A' + str(i)].border, ws['B' + str(i)].border, ws['C' + str(i)].border, ws['D' + str(i)].border, ws[
                    'E' + str(i)].border = linestyle, linestyle, linestyle, linestyle, linestyle
        return ws

    def kpi_chart(self, ws):
        chart = PieChart()
        labels = Reference(ws, min_col=5, min_row=12, max_row=14)
        data = Reference(ws, min_col=7, min_row=12, max_row=14) # 해당 셀의 데이터로 pie chart가 그려진다.
        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = "KEY PERFORMANCE CHART"
        chart.height = 6.8
        chart.width = 7.5
        slice = DataPoint(idx=0, explosion=20)
        chart.series[0].data_points = [slice]
        ws.add_chart(chart, "A7")
        return ws

    def trend_chart(self, ws: object, length: int): # trend 데이터를 그린다.
        chart = LineChart()
        if self.div == "Validation":
            P_datas = Reference(ws, min_col=4, min_row=34, max_row=35 + length)
            chart.title = "Validation TREND"
            chart.width = 18
        elif self.div == "Trend":
            P_datas = Reference(ws, min_col=4, min_row=34, max_row=35 + length)
            chart.title = "Process TREND"
            chart.width = 17
        chart.add_data(P_datas)
        chart.height = 6.8
        ws.add_chart(chart, "A20")
        return ws

    def make_excel_report(self):
        house_tag = self.dbmodule.selectHousetagByAnalyzertag(self.ana_tag)  # 단일 값으로 반환 된다.
        datas = self.dbmodule.dataForReport(self.start_dt, self.end_dt, self.div, self.ana_tag)  # ana_tag는 div가 Whole인 경우에는 필요 없다.

        (headerFill, header2Fill, tableFill, thin_border, top_bottom_border) = self.design_for_report()  # 셀 디자인만
        (availability_rate, check_rate, break_rate) = self.kpimodule.calculate_kpi()  # 각종 필요한 kpi요소
        if self.bottle_tag != None:
            reproducibility_rate = self.kpimodule.calculate_reproducibility(self.ana_tag, self.start_dt, self.end_dt, self.bottle_tag)  # 재현성 값
        else:  # trend의 경우에는 validation데이터가 아니므로 재현성 자체가 의미 없음.
            reproducibility_rate = 'None'
        (mtbf, mttr, mttf) = self.kpimodule.calculate_mean_break_time()  # 필요한 mtbf요소
        write_wb = openpyxl.Workbook()
        ws = write_wb.active

        ws.merge_cells('A1:F4')  # title용도 병합
        ws['A1'].fill = headerFill  # 헤더 부분에 채우기색상 적용
        ws = self.title_cell(ws, self.div, "TITLE")  # 보고서 맨 상단의 header 지정

        ws.merge_cells('G1:H2')
        ws['G1'].fill = header2Fill  # 오른쪽 헤더 부분에 채우기색상2 적용
        ws = self.cell_method(ws, 'G1', 'AMADAS', 11)

        ws.merge_cells('A5:D6')
        ws = self.cell_method(ws, 'A5', 'PERFORMANCE HISTORY CHART', 11, True)

        label = ['House', 'Analyzer', 'Period Start', 'Period End', 'Operator']
        value = [house_tag, self.ana_tag, self.start_dt, self.end_dt, self.user_id]

        ws = self.basic_inform_cell(ws, 5, 10, label, value, thin_border)

        rate = ['Availability Rate', 'Checking Rate', 'Breakdown Rate',
                'Reproducibility', 'MTBF', 'MTTF', 'MTTR']
        values = [round(float(availability_rate), 2), round(float(check_rate), 2), round(float(break_rate), 2),
                  round(float(reproducibility_rate), 2), round(float(mtbf), 2), round(float(mttf), 2),
                  round(float(mttr), 2)]

        ws = self.basic_inform_cell(ws, 12, 19, rate, values, thin_border)

        ws.merge_cells('A32:H33')
        ws = self.data_array(ws, self.div, datas, top_bottom_border, tableFill)

        ws = self.title_cell(ws, self.div, "VALUE")  # 데이터 테이블의 header 지정
        ws = self.column_cell(ws, self.div, top_bottom_border)  # 데이터 테이블의 column명 지정

        ws = self.kpi_chart(ws)  # kpi 차트 추가
        ws = self.trend_chart(ws, len(datas))
        print("정상 출력")
        write_wb.save(self.file_path() + self.file_name + ".xlsx")  # 최종 파일 생성
        return "OK"