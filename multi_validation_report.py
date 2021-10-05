import sys, time, datetime, openpyxl
from lxml import etree

import numpy as np
from matplotlib import pyplot as plt
import matplotlib
import xlrd
from openpyxl.styles import *
from openpyxl.cell import Cell
from openpyxl.chart import (PieChart, LineChart, ProjectedPieChart, Reference)
from openpyxl.chart.series import DataPoint
from decimal import *
from pylab import title, figure, xlabel, ylabel, xticks, bar, legend, axis, savefig

from db_module import *
from kpi_module import *
from report_abstract import *


class Multi_Validation_Report(Report): # 다중노드의 경우 validation report는 이쪽으로 들어온다.
    ana_tag: str
    user_id: str
    start_dt: str
    end_dt: str
    file_name: str
    count: int
    div: str
    dbmodule: object
    kpimodule: object

    def __init__(self, ana_tag, user_id, start_dt, end_dt, file_name, div, bottle_tag=None):
        self.ana_tag = ana_tag
        self.user_id = user_id
        self.start_dt = start_dt
        self.end_dt = end_dt
        self.file_name = file_name
        self.div = div
        self.bottle_tag = bottle_tag
        self.dbmodule = withDB()
        self.taggs = self.dbmodule.selectValidationTagByTag(ana_tag)
        self.count = len(self.taggs)

        self.dbmodule = withDB()
        self.kpimodule = KPI_module(start_dt, end_dt, div, ana_tag)

    def basic_inform_cell(self, ws: object, start_idx: int, end_idx: int, title: list, value: list, style: object): # 셀 그리는 기능
        for i in range(start_idx, end_idx):
            ws.merge_cells('F' + str(i) + ':G' + str(i))
            ws = self.cell_method(ws, 'F' + str(i), title[i - start_idx], 11, True)
            ws.merge_cells('H' + str(i) + ':I' + str(i))
            ws = self.cell_method(ws, 'H' + str(i), value[i - start_idx], 11)
            ws['F' + str(i)].border, ws['G' + str(i)].border, ws['H' + str(i)].border, ws[
                'I' + str(i)].border = style, style, style, style
        return ws

    def data_preprocess(self, datas: list, div: str): # validation 데이터를 가져와서 list별로 분류하는 기능.
        v_type, validation_value, start_time, check_time, result, user_id = [], [], [], [], [], []
        for i in range(0, self.count):
            v_type.append([])
            validation_value.append([])  # 3개 라면 [[],[],[]] 로 처음 배열 생성.
            start_time.append([])
            check_time.append([])
            result.append([])
            user_id.append([])
        for j in range(0, self.count):
            for k in range(0, len(datas)):
                if datas[k][0] == self.taggs[j][0]:
                    v_type[j].append(datas[k][1])
                    validation_value[j].append(round(float(datas[k][2]), 2))
                    start_time[j].append(str(datas[k][3])[2:16])
                    check_time[j].append(str(datas[k][4])[2:16])
                    result[j].append(datas[k][5])
                    user_id[j].append(datas[k][6])
        return (v_type, validation_value, start_time, check_time, result, user_id)

    def data_array(self, ws: object, div: str, datas: list, linestyle: object, fillstyle: object): # data를 반복문을 이용해서 파일에 쓴다.
        preprocessing_data = self.data_preprocess(datas, div)
        v_type, validation_value, start_time, check_time, result, user_id = preprocessing_data[0], preprocessing_data[
            1], preprocessing_data[2], preprocessing_data[3], preprocessing_data[4], preprocessing_data[5]
        row = len(validation_value[0])  # 태그 하나의 갯수가 곧 row줄 갯수가 된다.
        col_num = len(validation_value)  # 태그 갯수가 곧 column에 몇 개 추가 되어야 하는지를 나타낸다.
        cell_char = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
                     'W', 'X', 'Y', 'Z']

        for i in range(35, 35 + row):
            if i % 2 == 0:
                ws = self.cell_method(ws, 'A' + str(i), str(i - 35), 11)
                ws = self.cell_method(ws, 'B' + str(i), v_type[0][i - 35], 11)
                for k in range(0, col_num):
                    ws = self.cell_method(ws, cell_char[k] + str(i), validation_value[k][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num] + str(i), start_time[0][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num + 1] + str(i), check_time[0][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num + 2] + str(i), result[0][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num + 3] + str(i), user_id[0][i - 35], 11)

                ws['A' + str(i)].fill, ws['B' + str(i)].fill = fillstyle, fillstyle
                for m in range(1, col_num + 5):
                    ws[cell_char[m - 1] + str(i)].fill = fillstyle
            else:
                ws = self.cell_method(ws, 'A' + str(i), str(i - 35), 11)
                ws = self.cell_method(ws, 'B' + str(i), v_type[0][i - 35], 11)
                for k in range(0, col_num):
                    ws = self.cell_method(ws, cell_char[k] + str(i), validation_value[k][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num] + str(i), start_time[0][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num + 1] + str(i), check_time[0][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num + 2] + str(i), result[0][i - 35], 11)
                ws = self.cell_method(ws, cell_char[col_num + 3] + str(i), user_id[0][i - 35], 11)
            ws['A' + str(i)].border, ws['B' + str(i)].border = linestyle, linestyle
            for m in range(1, col_num + 5):
                ws[cell_char[m - 1] + str(i)].border = linestyle
        return ws

    def column_cell(self, ws: object, div: str, style: object): # data에 해당되는 column을 파일에 쓴다.
        valid_tag = self.taggs  # (('1630AI355A-VV',), ('1630AI355B-VV',), ('1630AI355C-VV',))
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                   'U', 'V', 'W', 'X', 'Y', 'Z']
        column_name = []
        width_no = []
        if div == "Validation":
            column_width = [5, 10, 15, 15, 10, 10]
            fixed_column_name = ['no', 'Type', 'Start Date', 'Check Date', 'Result', 'User']
            for i in range(0, 2 + len(valid_tag)):
                if i < 2:
                    column_name.append(fixed_column_name[i])
                    width_no.append(column_width[i])
                else:
                    column_name.append(valid_tag[i - 2][0])
                    width_no.append(15)
            for k in range(2 + len(valid_tag), len(valid_tag) + 6):
                column_name.append(fixed_column_name[k - len(valid_tag)])
                width_no.append(column_width[k - len(valid_tag)])

        elif div == "Trend":
            column_width = [5, 15, 15]
            fixed_column_name = ['no', 'Start Date', 'Check Date']
            for i in range(0, 1 + len(valid_tag)):
                if i < 1:
                    column_name.append(fixed_column_name[i])
                    width_no.append(column_width[i])
                else:
                    column_name.append(valid_tag[i - 1][0])
                    width_no.append(15)
            for k in range(1 + len(valid_tag), 3 + len(valid_tag)):
                column_name.append(fixed_column_name[k - len(valid_tag)])
                width_no.append(column_width[k - len(valid_tag)])
        for i in range(0, len(column_name)):
            ws[columns[i] + '34'] = column_name[i]
            ws[columns[i] + '34'].font = Font(name="맑은 고딕", size=11, bold=True)
            ws[columns[i] + '34'].alignment = Alignment(horizontal='center', vertical='center')
            ws[columns[i] + '34'].border = style
            ws.column_dimensions[columns[i]].width = width_no[i]
        return ws

    def kpi_chart(self, ws: object): # KPI 차트를
        chart = PieChart()
        label_min_col = 3
        data_min_col = 5
        labels = Reference(ws, min_col=(label_min_col + self.count), min_row=12, max_row=14) # 이 셀에 써진 값이 label로 나타난다.
        data = Reference(ws, min_col=(data_min_col + self.count), min_row=12, max_row=14) # 이 셀에 써진 값이 pie로 나타난다.
        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = "KEY PERFORMANCE CHART"
        chart.height = 6.8
        chart.width = 11
        slice = DataPoint(idx=0, explosion=5)
        chart.series[0].data_points = [slice]
        ws.add_chart(chart, "B7")
        return ws

    def trend_chart(self, ws: object, length: int):
        chart = LineChart()
        data_max_col = 2
        P_datas = Reference(ws, min_col=3, min_row=35, max_col=(data_max_col + self.count),
                            max_row=34 + int(length / self.count)) # 이 셀에 써진 값이 트렌드 그래프로 나타난다.
        chart.add_data(P_datas)
        chart.title = "VALUE TREND"
        chart.height = 6.8
        chart.width = 21
        ws.add_chart(chart, "A20")
        return ws

    def make_excel_report(self):
        house_tag = self.dbmodule.selectHousetagByAnalyzertag(self.ana_tag)
        datas = self.dbmodule.dataForReport(self.start_dt, self.end_dt, self.div, self.ana_tag)

        (headerFill, header2Fill, tableFill, thin_border, top_bottom_border) = self.design_for_report()  # 셀 디자인만 가져온다.
        (availability_rate, check_rate, break_rate) = self.kpimodule.calculate_kpi()  # 각종 필요한 kpi요소
        reproducibility_rate = self.kpimodule.calculate_reproducibility(self.ana_tag, self.start_dt, self.end_dt,
                                                                        self.bottle_tag)  # 재현성 값
        (mtbf, mttr, mttf) = self.kpimodule.calculate_mean_break_time()  # 필요한 mtbf요소

        write_wb = openpyxl.Workbook()
        ws = write_wb.active

        ws.merge_cells('A1:G4')
        ws['A1'].fill = headerFill
        ws = self.title_cell(ws, self.div, "TITLE")

        ws.merge_cells('H1:I2')
        ws['H1'].fill = header2Fill
        ws = self.cell_method(ws, 'H1', 'AMADAS', 11)

        ws.merge_cells('A5:E6')
        ws = self.cell_method(ws, 'A5', 'PERFORMANCE HISTORY CHART', 11, True)

        label = ['House', 'Analyzer', 'Period Start', 'Period End', 'Operator']
        value = [house_tag, self.ana_tag, self.start_dt, self.end_dt, self.user_id]
        ws = self.basic_inform_cell(ws, 5, 10, label, value, thin_border)

        rate = ['Availability Rate', 'Checking Rate', 'Breakdown Rate', 'Reproducibility',
                'MTBF', 'MTTF', 'MTTR']
        values = [round(float(availability_rate), 2), round(float(check_rate), 2), round(float(break_rate), 2),
                  round(float(reproducibility_rate), 2),
                  round(float(mtbf), 2), round(float(mttf), 2), round(float(mttr), 2)]

        ws = self.basic_inform_cell(ws, 12, 18, rate, values, thin_border)

        cell_char = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
        for i in range(1, 18):
            if self.count == i:
                ws.merge_cells('A32:' + cell_char[i] + '33')
                ws = self.data_array(ws, self.div, datas, top_bottom_border, tableFill)
                break
            else:
                pass

        ws = self.title_cell(ws, self.div, "VALUE")
        ws = self.column_cell(ws, self.div, top_bottom_border)

        ws = self.kpi_chart(ws)
        ws = self.trend_chart(ws, len(datas))

        print("정상 출력")
        write_wb.save(self.file_path() + self.file_name + ".xlsx")  # 최종 파일 생성
        return "OK"